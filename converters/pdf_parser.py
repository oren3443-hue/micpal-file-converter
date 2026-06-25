"""
Extract Michpal component-code → name mappings from a PDF file.
Looks for 3-digit codes followed by Hebrew text on the same line.
Uses PyMuPDF (fitz) which has no cryptography dependencies.
"""

import re
from typing import Dict

try:
    import fitz as _fitz
    _HAS_FITZ = True
except ImportError:
    _HAS_FITZ = False


# Pattern: optional "רכיב" prefix, then 3-digit code, separator, then description
_PATTERN = re.compile(
    r'(?:רכיב\s*)?(\d{3})\s*[-–—:]\s*([^\n\d]{2,60})',
    re.UNICODE,
)

# Table-like: "001    שכר בסיס"
_TABLE_PATTERN = re.compile(
    r'^(\d{3})\s{2,}(.{3,50})$',
    re.MULTILINE | re.UNICODE,
)


def _clean(name: str) -> str:
    return name.strip().strip('|').strip()


def extract_component_names(pdf_path: str) -> Dict[str, str]:
    """
    Return {code: name} dict from a PDF file.
    Silently returns empty dict on any failure.
    """
    result: Dict[str, str] = {}
    if not _HAS_FITZ:
        return result

    try:
        doc = _fitz.open(pdf_path)
        full_text = ''
        for page in doc:
            full_text += page.get_text() + '\n'
        doc.close()

        for m in _PATTERN.finditer(full_text):
            code, name = m.group(1), _clean(m.group(2))
            if name and 1 <= int(code) <= 999:
                result[code] = name

        for m in _TABLE_PATTERN.finditer(full_text):
            code, name = m.group(1), _clean(m.group(2))
            if code not in result and name and 1 <= int(code) <= 999:
                result[code] = name

    except Exception:
        pass

    return result


def extract_component_names_from_excel(xlsx_path: str) -> Dict[str, str]:
    """
    Return {code: name} dict from an Excel component-mapping file.
    Expects: column A = numeric code (1-252), column B = component name.
    Skips rows where name is empty or placeholder ('..........').
    Silently returns empty dict on any failure.
    """
    result: Dict[str, str] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

        # Prefer a sheet whose name contains 'רכיב' or 'שכר'
        ws = None
        for name in wb.sheetnames:
            if 'רכיב' in name or 'שכר' in name:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            try:
                code = str(int(row[0])).zfill(3)
                name = str(row[1]).strip() if row[1] else ''
                if name and not all(c in '._ ' for c in name):
                    result[code] = name
            except (ValueError, TypeError):
                continue

        wb.close()
    except Exception:
        pass

    return result
