"""
Excel attendance report parser.
Detects headers automatically and normalises column names.
"""

import openpyxl
from typing import List, Dict, Optional, Tuple


# Canonical column names mapped to Michpal report codes and field type (qty/price)
DEFAULT_COLUMN_MAPPING: Dict[str, Tuple[str, str]] = {
    'ימי עבודה בפועל':       ('124', 'qty'),
    'actual_work_days':      ('124', 'qty'),
    'payable_hours':         ('125', 'qty'),
    'שעות לתשלום':           ('125', 'qty'),
    '100%':                  ('301', 'qty'),
    '125%':                  ('307', 'qty'),
    'שעות 150 ללא לילה':     ('314', 'qty'),
    'שעות 150%':             ('314', 'qty'),
    'שעות לילה':             ('329', 'qty'),
    'שעות לילה ':            ('329', 'qty'),
    'הפסקה':                 ('330', 'qty'),
    'עלות נסיעות':           ('305', 'price'),
    'ימי חופש':              ('101', 'qty'),
    'vacation_days':         ('101', 'qty'),
    'ניצול חופשה':           ('101', 'qty'),
    'ימי מחלה לתשלום':      ('102', 'qty'),
    'sick_days_paid':        ('102', 'qty'),
    'ניצול מחלה':            ('102', 'qty'),
    'ימי מחלה':              ('102', 'qty'),
    'ימי מילואים':           ('104', 'qty'),
    'reserve_days':          ('104', 'qty'),
    'ניצול מילואים':         ('104', 'qty'),
    'צבירת חופשה':           ('111', 'qty'),
    'remaining_paid_leave_days': ('111', 'qty'),
}

# Column names that identify the payroll-system employee number
EMPLOYEE_NUM_ALIASES = {
    "מס' עובד במע' שכר",
    "מס' עובד בשכר",
    'employee_num_payroll',
    'payroll_employee_num',
    'מספר עובד שכר',
}


def _find_header_row(ws) -> Optional[int]:
    """Find the row that contains column headers (usually the first non-empty row)."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        non_empty = [c for c in row if c is not None]
        if len(non_empty) >= 3:
            return row_idx
    return 1


def parse_excel_file(filepath: str, column_mapping: Optional[Dict] = None) -> Dict:
    """
    Parse an Excel attendance report.

    Returns:
        {
            'employees': [
                {
                    'employee_num': str,
                    'id_num': str,
                    'first_name': str,
                    'last_name': str,
                    'components': [{'report_code': str, 'quantity': float, 'price': float}]
                },
                ...
            ],
            'columns': [{'header': str, 'report_code': str, 'field': str}],
            'year': int,
            'month': int,
        }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    headers = [cell.value for cell in ws[header_row]]

    mapping = column_mapping or DEFAULT_COLUMN_MAPPING

    # Map header positions to Michpal codes
    col_map = {}  # col_index -> (report_code, field)
    employee_col = None
    id_col = None
    first_name_col = None
    last_name_col = None

    for col_idx, header in enumerate(headers):
        if header is None:
            continue
        h = str(header).strip()

        if h in EMPLOYEE_NUM_ALIASES:
            employee_col = col_idx
        elif h in ('מספר עובד', 'employee_num', 'employee_id'):
            if employee_col is None:
                employee_col = col_idx
        elif h in ('מספר זהות', 'ת.ז.', 'id_num', 'identity'):
            id_col = col_idx
        elif h in ('שם פרטי', 'first_name', 'שם'):
            first_name_col = col_idx
        elif h in ('שם משפחה', 'last_name'):
            last_name_col = col_idx

        if h in mapping:
            col_map[col_idx] = mapping[h]

    if employee_col is None:
        raise ValueError("לא נמצאה עמודת מספר עובד בקובץ האקסל. "
                         "יש להוסיף עמודה בשם: 'מס' עובד במע' שכר'")

    employees = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        emp_val = row[employee_col] if employee_col < len(row) else None
        if emp_val is None:
            continue
        try:
            emp_num = str(int(emp_val))
        except (ValueError, TypeError):
            emp_num = str(emp_val).strip()
        if not emp_num or emp_num == '0':
            continue

        id_num = '000000000'
        if id_col is not None and id_col < len(row) and row[id_col]:
            id_num = str(row[id_col]).strip().zfill(9)[:9]

        first_name = ''
        if first_name_col is not None and first_name_col < len(row):
            first_name = str(row[first_name_col] or '')

        last_name = ''
        if last_name_col is not None and last_name_col < len(row):
            last_name = str(row[last_name_col] or '')

        components = []
        for col_idx, (report_code, field) in col_map.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            try:
                num = float(val)
            except (ValueError, TypeError):
                continue
            if num == 0:
                continue

            qty = num if field == 'qty' else 0.0
            price = num if field == 'price' else 0.0
            components.append({
                'report_code': report_code,
                'quantity': qty,
                'price': price,
            })

        if components:
            employees.append({
                'employee_num': emp_num,
                'id_num': id_num,
                'first_name': first_name,
                'last_name': last_name,
                'components': components,
            })

    # Build column info for UI display
    columns = []
    for col_idx, (report_code, field) in col_map.items():
        header = str(headers[col_idx]).strip() if col_idx < len(headers) else ''
        columns.append({
            'header': header,
            'report_code': report_code,
            'field': field,
            'col_idx': col_idx,
        })

    return {
        'employees': employees,
        'columns': columns,
        'year': None,
        'month': None,
    }
