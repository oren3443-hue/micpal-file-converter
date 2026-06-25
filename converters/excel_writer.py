"""
Write a formatted RTL Hebrew Excel file from Michpal records.
"""

from collections import defaultdict
from typing import List, Dict, Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side,
)
from openpyxl.utils import get_column_letter

from .michpal_parser import MichpalRecord, BUILTIN_CODE_NAMES


# Colour palette
CLR_HEADER_BG   = '1F4E79'  # dark blue
CLR_HEADER_FG   = 'FFFFFF'
CLR_ROW_ALT     = 'DDEEFF'
CLR_ROW_WHITE   = 'FFFFFF'
CLR_SUBHEADER   = '2E75B6'  # medium blue
CLR_BORDER      = 'AAAAAA'


def _border(style='thin'):
    s = Side(style=style, color=CLR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row, col, value, wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name='Arial', bold=True, size=10, color=CLR_HEADER_FG)
    cell.fill = PatternFill(start_color=CLR_HEADER_BG, end_color=CLR_HEADER_BG, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    cell.border = _border()
    return cell


def _data_cell(ws, row, col, value, is_alt=False, number_format=None, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name='Arial', size=10, bold=bold)
    bg = CLR_ROW_ALT if is_alt else CLR_ROW_WHITE
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _border()
    if number_format:
        cell.number_format = number_format
    return cell


def _component_code(report_code: str) -> str:
    """Human-readable component number for a raw QDIV report code.
    Salary components 1-252 are stored as 301-552 in the file (+300).
    Attendance/built-in codes (100-131, 900s) are stored as-is.
    """
    code = int(report_code)
    if 300 < code <= 552:
        return str(code - 300).zfill(3)
    return report_code


def _name_for_code(report_code: str, component_names: Dict) -> str:
    """
    Resolve a display name for a raw QDIV report code.

    Priority:
    1. Attendance/built-in codes (BUILTIN_CODE_NAMES) — matched on raw code.
    2. Salary components (QDIV 301-552 = user components 001-252) — looked up
       by component number (code-300) in component_names.
    3. Fallback: 'רכיב {component_number}'.

    This avoids the collision where attendance code 124 and salary component 124
    (stored as QDIV 424) would otherwise map to the same display key.
    """
    # Attendance / built-in: raw code is the canonical key
    if report_code in BUILTIN_CODE_NAMES:
        return BUILTIN_CODE_NAMES[report_code]
    code = int(report_code)
    # Salary component: subtract offset to get component number
    if 300 < code <= 552:
        comp_key = str(code - 300).zfill(3)
        return component_names.get(comp_key, f'רכיב {comp_key}')
    # Other codes: try direct lookup, then fallback
    return component_names.get(report_code, f'קוד {report_code}')


def write_excel_from_michpal(
    records: List[MichpalRecord],
    output_path: str,
    component_names: Optional[Dict[str, str]] = None,
) -> None:
    """
    Create a formatted RTL Excel from Michpal records.
    One row per employee, one (or two) columns per report code.
    """
    if not records:
        raise ValueError("אין רשומות לכתיבה")

    comp_names = component_names or {}

    # Group by employee; use RAW report_code as column key to avoid collisions
    # between attendance codes (e.g. 124) and salary components (e.g. 424 = comp 124).
    employees: Dict[str, Dict] = defaultdict(lambda: {
        'id_num': '',
        'codes': {},  # raw report_code -> {'qty': float, 'price': float}
        'company': '',
        'yymm': '',
    })

    all_codes = sorted(set(r.report_code for r in records))

    for rec in records:
        if rec.record_code == '9':
            continue
        emp = rec.employee_num
        code_key = rec.report_code
        employees[emp]['id_num'] = rec.id_num
        employees[emp]['company'] = rec.company
        employees[emp]['yymm'] = rec.yymm
        if code_key not in employees[emp]['codes']:
            employees[emp]['codes'][code_key] = {'qty': 0.0, 'price': 0.0}
        employees[emp]['codes'][code_key]['qty'] += rec.quantity
        if rec.price != 0:
            employees[emp]['codes'][code_key]['price'] = rec.price

    # Determine which codes have non-zero prices
    codes_with_price: set = set()
    for emp_data in employees.values():
        for code, vals in emp_data['codes'].items():
            if vals['price'] != 0:
                codes_with_price.add(code)

    # Build column definitions: (raw_code, field, label)
    col_defs = []
    for code in all_codes:
        code_name = _name_for_code(code, comp_names)
        col_defs.append((code, 'qty', f'{code_name}\n(כמות)'))
        if code in codes_with_price:
            col_defs.append((code, 'price', f'{code_name}\n(מחיר)'))

    # ── Workbook setup ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'נתוני מיכפל'
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = True

    # ── Period header row ──
    first = records[0]
    period_label = f"נתוני שכר {first.month:02d}/{first.year}  |  חברה {first.company}"
    total_cols = 2 + len(col_defs)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    cell = ws.cell(row=1, column=1, value=period_label)
    cell.font = Font(name='Arial', bold=True, size=13, color=CLR_HEADER_FG)
    cell.fill = PatternFill(start_color=CLR_SUBHEADER, end_color=CLR_SUBHEADER, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # ── Column headers (row 2) ──
    ws.row_dimensions[2].height = 50
    _header_cell(ws, 2, 1, 'מספר עובד')
    _header_cell(ws, 2, 2, 'מספר זהות')
    for c_idx, (code, field, label) in enumerate(col_defs, 3):
        _header_cell(ws, 2, c_idx, label)

    # ── Data rows ──
    sorted_employees = sorted(employees.items(),
                              key=lambda kv: int(kv[0]) if kv[0].isdigit() else 0)
    for row_offset, (emp_num, emp_data) in enumerate(sorted_employees):
        row = 3 + row_offset
        is_alt = row_offset % 2 == 0

        _data_cell(ws, row, 1, emp_num, is_alt)
        _data_cell(ws, row, 2, emp_data['id_num'], is_alt)

        for c_idx, (code, field, _label) in enumerate(col_defs, 3):
            vals = emp_data['codes'].get(code, {'qty': 0.0, 'price': 0.0})
            val = vals[field]
            display_val = round(val, 2) if val != 0 else None
            _data_cell(ws, row, c_idx, display_val, is_alt, number_format='#,##0.00')

    last_data_row = 2 + len(sorted_employees)

    # ── AutoFilter on header row ──
    ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{last_data_row}"

    # ── Freeze header rows ──
    ws.freeze_panes = 'A3'

    # ── Column widths ──
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 13
    for c_idx in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 16

    # ── Summary sheet ──
    ws2 = wb.create_sheet('סיכום')
    ws2.sheet_view.rightToLeft = True
    meta = [
        ('מספר חברה', first.company),
        ('שנה', first.year),
        ('חודש', first.month),
        ('מספר עובדים', len(employees)),
        ('מספר רשומות', len(records)),
        ('סמלי דיווח', ', '.join(_component_code(c) for c in all_codes)),
    ]
    for i, (k, v) in enumerate(meta, 1):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True, name='Arial')
        ws2.cell(row=i, column=2, value=v).font = Font(name='Arial')
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 40

    wb.save(output_path)
